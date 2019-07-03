#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : do_excel.py
# @Author: LILIANG
# @Date  : 2019/6/21
# @Desc  :  test


import openpyxl


class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname
        self.workbook=openpyxl.load_workbook(filename)
        self.sheet=self.workbook[sheetname]



    def get_cases(self):

        max_row=self.sheet.max_row
        cases=[]
        for r in range(2,max_row+1):
            case=Case()
            case.case_id   =self.sheet.cell(r,column=1).value
            case.title=self.sheet.cell(r,column=2).value
            case.method = self.sheet.cell(row=r, column=3).value
            case.url = self.sheet.cell(row=r, column=4).value
            case.data = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value  # sql
            cases.append(case)

        self.workbook.close()
        return cases


    def writeResult(self,row,actual,result):
        sheet=self.workbook[self.sheetname]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        self.workbook.save(filename=self.filename)
        self.workbook.close()





